from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Any

import pandas as pd

try:
    from openpyxl.chart import BarChart, LineChart, Reference
except ImportError:  # Charts are optional. The workbook will still be written.
    BarChart = LineChart = Reference = None

from map_country import (
    PUBLICATION_LOOKBACK_YEARS,
    FILTER_BY_PUBLICATION_DATE,
    KEEP_ROWS_WITH_MISSING_PUBLICATION_DATE,
    EXCLUDE_PROVISIONAL_RECORDS,
    RELEVANCY_KEYWORDS,
    TITLE_CASE_TEXT_FIELDS,
    USE_APPLICATION_COUNTRY_WHEN_PRIORITY_COUNTRY_IS_WO,
    DUPLICATE_APPLICATION_STATUS_POLICY,
    UNIQUE_FAMILY_SHEET_NAME,
    USER_FACING_COLUMNS_TO_HIDE,
    TEXT_COLUMNS_TO_FORMAT,
    REQUIRED_COLUMNS,
    OPTIONAL_COLUMNS,
    FORWARD_CITATION_COLUMN_CANDIDATES,
    EP_MEMBER_OR_VALIDATION_COUNTRY_CODES,
    FIXED_COUNTRY_PRIORITY,
    APPLICATION_STATUS_SORT_RANK,
    DEAD_ALIVE_SORT_RANK,
    COUNTRY_MAP,
    APPLICATION_STATUS_PAIRS,
)


# ---------------------------------------------------------------------------
# Generic helpers
# ---------------------------------------------------------------------------

def clean_identifier_value(value: Any) -> str | pd.NA:
    """Return a stripped identifier while preserving leading zeroes and alphanumeric IDs."""
    if pd.isna(value):
        return pd.NA
    text = str(value).strip()
    if text in {"", "-", "--", "nan", "NaN", "None", "NONE", "null", "NULL"}:
        return pd.NA
    if text.endswith(".0") and text[:-2].isdigit():
        # Excel/pandas can sometimes expose integer-like identifiers as 12345.0.
        text = text[:-2]
    return text


def clean_identifier_series(series: pd.Series) -> pd.Series:
    return series.map(clean_identifier_value).astype("string")


def normalize_patent_number(value: Any) -> str:
    cleaned = clean_identifier_value(value)
    if pd.isna(cleaned):
        return ""
    return re.sub(r"[^A-Za-z0-9]", "", str(cleaned)).upper()


def first_nonblank_token(value: Any) -> str:
    cleaned = clean_identifier_value(value)
    if pd.isna(cleaned):
        return ""
    parts = re.split(r"[|;,]", str(cleaned))
    for part in parts:
        token = part.strip()
        if token and token != "-":
            return token
    return ""


def extract_country_code_from_identifier(value: Any) -> str | pd.NA:
    token = normalize_patent_number(first_nonblank_token(value))
    match = re.match(r"^([A-Z]{2})", token)
    return match.group(1) if match else pd.NA


def normalize_string_for_matching(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip().lower()


def add_note(notes: list[dict[str, Any]], category: str, message: str, count: Any = "") -> None:
    notes.append({"Category": category, "Message": message, "Count": count})


def blank_to_na(series: pd.Series) -> pd.Series:
    """Convert blank-like values to pandas NA."""
    text = series.astype("string").str.strip()
    return text.mask(
        text.isna()
        | text.eq("")
        | text.eq("-")
        | text.str.lower().isin(["nan", "none", "null"])
    )


def normalize_column_name(value: Any) -> str:
    """Normalize a column name for tolerant matching."""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


# ---------------------------------------------------------------------------
# Mapping helpers
# ---------------------------------------------------------------------------

def build_application_status_map() -> tuple[dict[str, str], pd.DataFrame]:
    mapping: dict[str, str] = {}
    seen: dict[str, list[str]] = defaultdict(list)

    for raw_key, raw_value in APPLICATION_STATUS_PAIRS:
        key = str(raw_key).strip().upper()
        value = str(raw_value).strip()
        seen[key].append(value)

        if DUPLICATE_APPLICATION_STATUS_POLICY == "first" and key in mapping:
            continue
        mapping[key] = value

    duplicate_rows: list[dict[str, Any]] = []
    for key, values in seen.items():
        if len(values) > 1:
            unique_values = list(dict.fromkeys(values))
            duplicate_rows.append(
                {
                    "Application Status Code": key,
                    "Occurrences": len(values),
                    "Values Seen": " | ".join(values),
                    "Unique Values": " | ".join(unique_values),
                    "Conflict": len(set(values)) > 1,
                    "Kept Value": mapping[key],
                    "Policy": DUPLICATE_APPLICATION_STATUS_POLICY,
                }
            )

    return mapping, pd.DataFrame(duplicate_rows)


def extract_application_status_code(value: Any, status_map: dict[str, str]) -> str | pd.NA:
    pub = normalize_patent_number(value)
    if len(pub) < 3:
        return pd.NA

    country_code = pub[:2]
    max_suffix_length = max(len(code) - 2 for code in status_map)
    for suffix_length in range(min(max_suffix_length, max(len(pub) - 2, 0)), 0, -1):
        candidate = country_code + pub[-suffix_length:]
        if candidate in status_map:
            return candidate

    # Fallback only records the best-looking kind code. It may remain unmapped.
    fallback_match = re.search(r"([A-Z]{1,3}\d{0,2})$", pub)
    if fallback_match:
        return country_code + fallback_match.group(1)
    return pd.NA


# ---------------------------------------------------------------------------
# Data preparation
# ---------------------------------------------------------------------------

def validate_and_add_optional_columns(data: pd.DataFrame, notes: list[dict[str, Any]]) -> pd.DataFrame:
    missing_required = [column for column in REQUIRED_COLUMNS if column not in data.columns]
    if missing_required:
        raise ValueError(f"Missing required columns: {missing_required}")

    for column in OPTIONAL_COLUMNS:
        if column not in data.columns:
            data[column] = pd.NA
            add_note(notes, "Missing optional column", f"Created blank column: {column}")

    return data


def clean_text_columns(data: pd.DataFrame) -> pd.DataFrame:
    for column in TEXT_COLUMNS_TO_FORMAT:
        if column not in data.columns:
            continue
        text = data[column].astype("string").str.strip()
        if TITLE_CASE_TEXT_FIELDS:
            text = text.str.title()
        data[column] = text
    return data


def fill_current_assignee(data: pd.DataFrame, notes: list[dict[str, Any]]) -> pd.DataFrame:
    """Fill Current Assignees using Current Assignees, Original Assignees, then Applicants."""
    before_blank = int(blank_to_na(data["Current Assignees"]).isna().sum())

    current_assignee = blank_to_na(data["Current Assignees"])
    original_assignee = blank_to_na(data["Original Assignees"])
    applicants = blank_to_na(data["Applicants"])

    data["Current Assignees"] = current_assignee.fillna(original_assignee).fillna(applicants)

    after_blank = int(blank_to_na(data["Current Assignees"]).isna().sum())

    add_note(
        notes,
        "Current Assignees fallback",
        "Filled Current Assignees using Current Assignees, Original Assignees, then Applicants",
        f"{before_blank - after_blank} filled / {after_blank} still blank",
    )

    return data


def prepare_identifiers(data: pd.DataFrame) -> pd.DataFrame:
    identifier_columns = [
        "Publication Number",
        "Application Number",
        "INPADOC Family ID",
        "INPADOC Family Members (Beta)",
        "Simple Family ID",
        "Priority Number",
    ]
    for column in identifier_columns:
        if column in data.columns:
            data[column] = clean_identifier_series(data[column])

    data["Application Number"] = data["Application Number"].fillna(data["Publication Number"])
    data["INPADOC Family ID"] = data["INPADOC Family ID"].fillna(data["Publication Number"])
    data["INPADOC Family Members (Beta)"] = data["INPADOC Family Members (Beta)"].fillna(data["Publication Number"])
    data["Priority Number"] = data["Priority Number"].fillna(data["Application Number"])
    return data


def prepare_dates_and_years(data: pd.DataFrame) -> pd.DataFrame:
    for column in ["Publication Date", "Filing Date", "Priority Date"]:
        data[column] = pd.to_datetime(data[column], errors="coerce")

    data["Priority Date"] = data["Priority Date"].fillna(data["Filing Date"])
    data["Filing Date"] = data["Filing Date"].fillna(data["Priority Date"])
    
    data["Publication Year"] = data["Publication Date"].dt.year.astype("Int64")
    data["Application Year"] = data["Filing Date"].dt.year.astype("Int64")
    return data


def apply_publication_date_filter(data: pd.DataFrame, notes: list[dict[str, Any]]) -> pd.DataFrame:
    if not FILTER_BY_PUBLICATION_DATE:
        add_note(notes, "Publication date filter", "Skipped because FILTER_BY_PUBLICATION_DATE is False")
        return data

    cutoff = pd.Timestamp.today().normalize() - pd.DateOffset(years=PUBLICATION_LOOKBACK_YEARS)
    before = len(data)
    mask = data["Publication Date"] >= cutoff
    if KEEP_ROWS_WITH_MISSING_PUBLICATION_DATE:
        mask = mask | data["Publication Date"].isna()
    filtered = data.loc[mask].copy()
    add_note(
        notes,
        "Publication date filter",
        f"Kept records with Publication Date on or after {cutoff.date()}",
        f"{len(filtered)} kept / {before - len(filtered)} removed",
    )
    return filtered


# ---------------------------------------------------------------------------
# Field creation
# ---------------------------------------------------------------------------

def add_application_status(data: pd.DataFrame, status_map: dict[str, str], notes: list[dict[str, Any]]) -> pd.DataFrame:
    data["Application Status Code"] = data["Publication Number"].map(
        lambda value: extract_application_status_code(value, status_map)
    ).astype("string")
    data["Application Status"] = data["Application Status Code"].map(status_map).fillna("Unknown")

    unknown_count = int((data["Application Status"] == "Unknown").sum())
    if unknown_count:
        add_note(notes, "Application Status", "Application Status Code was not found in mapping", unknown_count)
    return data


def add_country_fields(data: pd.DataFrame) -> pd.DataFrame:
    data["Country Code"] = data["Publication Number"].map(extract_country_code_from_identifier).astype("string")
    data["Country Name"] = data["Country Code"].map(COUNTRY_MAP).fillna("Unknown")

    data["Application Country Code"] = data["Application Number"].map(extract_country_code_from_identifier).astype("string")
    priority_number = data["Priority Number"].astype("string")
    priority_number_clean = priority_number.fillna("")

    pct_priority_mask = priority_number_clean.str.startswith("PCT")
    priority_country_code_raw = priority_number.map(extract_country_code_from_identifier).astype("string")
    priority_country_code_raw = priority_country_code_raw.mask(
        pct_priority_mask,
        priority_number_clean.str[3:5].where(priority_number_clean.str.len() >= 5, pd.NA),
    )
    data["Priority Country Code Raw"] = priority_country_code_raw
    data["Priority Country Code"] = data["Priority Country Code Raw"]

    priority_country_column = data["Priority Country"].astype("string") if "Priority Country" in data.columns else pd.Series(pd.NA, index=data.index, dtype="string")
    priority_country_column = priority_country_column.str.strip().str.upper()

    priority_country_code = data["Priority Country Code"].copy()
    priority_country_mask = priority_country_column.isin(["WO", "IB"])
    priority_country_code = priority_country_code.mask(priority_country_mask, data["Country Code"])

    still_wo_ib_mask = priority_country_code.isin(["WO", "IB"])
    application_number = data["Application Number"].astype("string").fillna("")
    application_is_pct = application_number.str.startswith("PCT")
    application_country_from_pct = application_number.str[3:5].where(application_number.str.len() >= 5, pd.NA)
    application_country_from_non_pct = application_number.str[:2].where(application_number.str.len() >= 2, pd.NA)
    application_country_fallback = application_country_from_non_pct.mask(application_is_pct, application_country_from_pct)
    priority_country_code = priority_country_code.mask(still_wo_ib_mask, application_country_fallback)

    if USE_APPLICATION_COUNTRY_WHEN_PRIORITY_COUNTRY_IS_WO:
        mask = (
            priority_country_code.eq("WO")
            & data["Application Country Code"].notna()
            & ~data["Application Country Code"].eq("WO")
        )
        priority_country_code = priority_country_code.mask(mask, data["Application Country Code"])

    data["Priority Country Code"] = priority_country_code.astype("string")

    data["Priority Country/Region"] = data["Priority Country Code"]
    data["Priority Country/Region Full"] = data["Priority Country Code"].map(COUNTRY_MAP).fillna("Unknown")
    return data


def is_design_publication_number(value: Any) -> bool:
    """Return True for publication numbers starting with USD or ending with S/S1/S2."""
    pub = normalize_patent_number(value)
    if not pub:
        return False
    return pub.startswith("USD") or pub.endswith(("S", "S1", "S2"))


def derive_patent_type_from_publication(publication_number: Any, existing_patent_type: Any = pd.NA) -> str:
    """Derive Patent Type so the raw upload does not need this column."""
    if is_design_publication_number(publication_number):
        return "Designs"

    cleaned_existing = clean_identifier_value(existing_patent_type)
    if not pd.isna(cleaned_existing):
        return str(cleaned_existing).strip()

    return "Patents"


def add_patent_type(data: pd.DataFrame, notes: list[dict[str, Any]]) -> pd.DataFrame:
    original_present = "Patent Type" in data.columns
    if not original_present:
        data["Patent Type"] = pd.NA

    original_values = data["Patent Type"].copy()
    design_mask = data["Publication Number"].map(is_design_publication_number)
    blank_mask = original_values.map(clean_identifier_value).isna()

    data["Patent Type"] = data.apply(
        lambda row: derive_patent_type_from_publication(row["Publication Number"], row.get("Patent Type", pd.NA)),
        axis=1,
    )

    data["Patent Type Source"] = "Input Patent Type"
    data.loc[design_mask, "Patent Type Source"] = "Derived from Publication Number design rule"
    data.loc[~design_mask & blank_mask, "Patent Type Source"] = "Defaulted to Patents because source was blank/missing"

    add_note(
        notes,
        "Patent Type",
        "Derived designs from publication numbers starting with USD or ending with S/S1/S2",
        int(design_mask.sum()),
    )
    if not original_present:
        add_note(notes, "Patent Type", "Patent Type column was missing in input; created automatically")
    if int((~design_mask & blank_mask).sum()):
        add_note(
            notes,
            "Patent Type",
            "Blank/missing non-design Patent Type values defaulted to Patents for 240-month rule",
            int((~design_mask & blank_mask).sum()),
        )

    return data


def create_ar_and_relevancy_fields(data: pd.DataFrame) -> pd.DataFrame:
    abstract_candidates = [
        "Abstract",
        "Abstract (Translated)(English)",
        "Abstract(Translated)(English)",
        "Abstract (English)",
        "Abstract(Original)",
    ]
    representative_claim_candidates = [
        "Representative Claim",
        "Representative Claims",
        "First Claim",
        "Independent Claim",
        "Claims",
        "Claim",
    ]

    abstract_col = next((column for column in abstract_candidates if column in data.columns), None)
    claim_col = next((column for column in representative_claim_candidates if column in data.columns), None)

    if abstract_col is None:
        abstract_text = pd.Series("", index=data.index, dtype="string")
    else:
        abstract_text = data[abstract_col].fillna("").astype("string").str.strip()

    if claim_col is None:
        claim_text = pd.Series("", index=data.index, dtype="string")
    else:
        claim_text = data[claim_col].fillna("").astype("string").str.strip()

    data["A+R"] = ("Abstract: " + abstract_text + "\nRepresentative Claim: " + claim_text).str.strip()

    if RELEVANCY_KEYWORDS:
        pattern = "|".join(re.escape(keyword) for keyword in RELEVANCY_KEYWORDS if keyword.strip())
        if pattern:
            relevant_mask = data["A+R"].str.contains(pattern, case=False, na=False, regex=True)
            data["A+R Relevancy"] = relevant_mask.map({True: "Relevant", False: "Not Relevant"})
        else:
            data["A+R Relevancy"] = "Review Required"
    else:
        existing = data["Relevancy Checking"].astype("string").str.strip()
        data["A+R Relevancy"] = existing.mask(existing.isna() | existing.eq(""), "Review Required")

    return data


# ---------------------------------------------------------------------------
# Alive/Dead logic
# ---------------------------------------------------------------------------

def projectx_status_to_dead_alive(status: Any) -> str | pd.NA:
    status = normalize_string_for_matching(status)

    if status in {"alive", "active"}:
        return "Active"

    if status in {"dead", "inactive"}:
        return "Inactive"

    return pd.NA


def add_dead_alive(data: pd.DataFrame) -> pd.DataFrame:
    today = pd.Timestamp.today().normalize()

    data["ProjectX Legal Status"] = data["Simple Legal Status"]

    status_frame = data.apply(
        lambda row: classify_dead_alive_for_row(row, today),
        axis=1
    )

    data["Dead_Alive"] = status_frame["Dead_Alive"]
    data["Dead_Alive Source"] = status_frame["Dead_Alive Source"]

    return data


def classify_dead_alive_for_row(row: pd.Series, today: pd.Timestamp) -> pd.Series:
    country_code = str(row.get("Country Code") or "").strip().upper()
    e_date = row.get("Priority Date")
    simple_status = row.get("ProjectX Legal Status")
    patent_type = normalize_string_for_matching(row.get("Patent Type"))

    if country_code == "WO":
        if pd.isna(e_date):
            return pd.Series({"Dead_Alive": "Review Required", "Dead_Alive Source": "WIPO 30-month rule; missing E-date"})
        if e_date > today - pd.DateOffset(months=30):
            return pd.Series({"Dead_Alive": "Active", "Dead_Alive Source": "WIPO 30-month rule"})
        return pd.Series({"Dead_Alive": "Inactive", "Dead_Alive Source": "WIPO 30-month rule"})

    projectx_value = projectx_status_to_dead_alive(simple_status)
    if not pd.isna(projectx_value):
        return pd.Series({"Dead_Alive": projectx_value, "Dead_Alive Source": "ProjectX Legal Status"})

    is_design = "design" in patent_type
    is_utility = (
        not is_design
        and (
            patent_type in {"utility models", "utility model", "applications", "application", "patents", "patent", "undetermined", ""}
            or "utility" in patent_type
        )
    )

    if pd.isna(e_date):
        return pd.Series({"Dead_Alive": "Review Required", "Dead_Alive Source": "Missing E-date"})

    if is_design:
        status = "Active" if e_date > today - pd.DateOffset(months=180) else "Inactive"
        return pd.Series({"Dead_Alive": status, "Dead_Alive Source": "Design 180-month E-date rule"})

    if is_utility:
        status = "Active" if e_date > today - pd.DateOffset(months=240) else "Inactive"
        return pd.Series({"Dead_Alive": status, "Dead_Alive Source": "Utility 240-month E-date rule"})

    return pd.Series({"Dead_Alive": "Review Required", "Dead_Alive Source": "Unmapped patent type/status"})

# ---------------------------------------------------------------------------
# Forward-citation counting
# ---------------------------------------------------------------------------

def find_forward_citation_column(data: pd.DataFrame) -> str | None:
    """Return the raw export column that contains pipe-separated forward citations."""
    normalized_to_original = {normalize_column_name(column): column for column in data.columns}

    for candidate in FORWARD_CITATION_COLUMN_CANDIDATES:
        normalized_candidate = normalize_column_name(candidate)
        if normalized_candidate in normalized_to_original:
            return normalized_to_original[normalized_candidate]

    # Fallback for PatSnap/export variants with slightly different labels.
    for column in data.columns:
        normalized = normalize_column_name(column)
        if "count" in normalized:
            continue
        has_forward = "forward" in normalized
        has_citation = "citation" in normalized or "cited" in normalized
        has_patent_hint = "patent" in normalized or "publication" in normalized or "number" in normalized
        if has_forward and has_citation and has_patent_hint:
            return column

    return None


def split_unique_forward_citations(value: Any) -> list[str]:
    """Split a pipe-separated citation cell and return unique normalized patent numbers."""
    cleaned = clean_identifier_value(value)
    if pd.isna(cleaned):
        return []

    unique_values: dict[str, None] = {}
    for part in str(cleaned).split("|"):
        patent_number = normalize_patent_number(part)
        if patent_number:
            unique_values[patent_number] = None
    return list(unique_values.keys())


def add_forward_citation_counts(data: pd.DataFrame, notes: list[dict[str, Any]]) -> pd.DataFrame:
    """Create Count of Cited by Patents from a pipe-separated forward-citation column."""
    forward_column = find_forward_citation_column(data)

    if forward_column:
        unique_lists = data[forward_column].map(split_unique_forward_citations)
        data["Forward Citation Unique Patent Numbers"] = unique_lists.map(lambda values: " | ".join(values))
        data["Count of Cited by Patents"] = unique_lists.map(len).astype(int)
        data["Count of Cited by Patents Source"] = f"Unique count from {forward_column}"
        add_note(
            notes,
            "Forward citations",
            f"Calculated Count of Cited by Patents from pipe-separated values in '{forward_column}'",
            f"{int((data['Count of Cited by Patents'] > 0).sum())} rows with citations",
        )
        return data

    data["Count of Cited by Patents"] = pd.to_numeric(
        data["Count of Cited by Patents"], errors="coerce"
    ).fillna(0).astype(int)
    data["Count of Cited by Patents Source"] = "Existing Count of Cited by Patents column"
    add_note(
        notes,
        "Forward citations",
        "No forward-citation patent-number column found; kept existing Count of Cited by Patents values",
    )
    return data


# ---------------------------------------------------------------------------
# Sorting, dedupe, and summaries
# ---------------------------------------------------------------------------

def build_country_priority(data: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, int]]:
    counts = data["Country Code"].dropna().astype(str).value_counts()
    country_codes = set(counts.index)

    ordered_codes: list[str] = []
    for code in FIXED_COUNTRY_PRIORITY:
        if code in country_codes:
            ordered_codes.append(code)

    ep_bucket = [
        code
        for code in EP_MEMBER_OR_VALIDATION_COUNTRY_CODES
        if code in country_codes and code not in ordered_codes
    ]
    ep_bucket.sort(key=lambda code: (-int(counts.get(code, 0)), code))
    ordered_codes.extend(ep_bucket)

    remaining = [code for code in country_codes if code not in set(ordered_codes)]
    remaining.sort(key=lambda code: (-int(counts.get(code, 0)), code))
    ordered_codes.extend(remaining)

    rank_map = {code: rank for rank, code in enumerate(ordered_codes, start=1)}
    data["Country Priority Rank"] = data["Country Code"].map(rank_map).fillna(len(rank_map) + 1).astype(int)
    return data, rank_map


def add_sort_columns(data: pd.DataFrame) -> pd.DataFrame:
    data["Application Status Sort Rank"] = data["Application Status"].map(APPLICATION_STATUS_SORT_RANK).fillna(9).astype(int)
    data["Dead_Alive Sort Rank"] = data["Dead_Alive"].map(DEAD_ALIVE_SORT_RANK).fillna(9).astype(int)
    return data


def sort_for_application_dedupe(data: pd.DataFrame) -> pd.DataFrame:
    return data.sort_values(
        by=["Application Status Sort Rank", "Publication Date", "Country Priority Rank", "Dead_Alive Sort Rank"],
        ascending=[True, False, True, True],
        na_position="last",
    ).copy()


def sort_for_family_dedupe(data: pd.DataFrame) -> pd.DataFrame:
    return data.sort_values(
        by=["Country Priority Rank", "Dead_Alive Sort Rank", "Application Status Sort Rank", "Publication Date"],
        ascending=[True, True, True, False],
        na_position="last",
    ).copy()


def drop_duplicates_on_nonblank(data: pd.DataFrame, column: str) -> tuple[pd.DataFrame, int, str]:
    if column not in data.columns:
        return data, 0, "Skipped because column is missing"

    values = clean_identifier_series(data[column])
    valid = values.notna() & values.ne("")
    duplicated = valid & values.duplicated(keep="first")
    removed = int(duplicated.sum())
    return data.loc[~duplicated].copy(), removed, "Completed"


def create_unique_applications(data: pd.DataFrame, notes: list[dict[str, Any]]) -> pd.DataFrame:
    before = len(data)
    unique_applications, removed, note = drop_duplicates_on_nonblank(data, "Application Number")
    add_note(notes, "Application dedupe", note, f"{removed} removed / {before} before / {len(unique_applications)} after")
    return unique_applications


def create_unique_patent_families(data: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    current = data.copy()
    rows: list[dict[str, Any]] = []
    for step_number, column in enumerate(["INPADOC Family ID", "Simple Family ID", "Priority Number"], start=1):
        before = len(current)
        current, removed, note = drop_duplicates_on_nonblank(current, column)
        rows.append(
            {
                "Step": step_number,
                "Deduplication Column": column,
                "Before": before,
                "Removed": removed,
                "After": len(current),
                "Note": note,
            }
        )
    return current, pd.DataFrame(rows)


def create_country_summary(data: pd.DataFrame, country_rank_map: dict[str, int]) -> pd.DataFrame:
    """Create country summary sorted by patent count for charts."""
    summary = (
        data.groupby(["Country Name"], dropna=False)
        .size()
        .reset_index(name="Patents")
        .rename(columns={"Country Name": "Country"})
    )
    return summary.sort_values(["Patents", "Country"], ascending=[False, True]).reset_index(drop=True)


def country_priority_group(country_code: Any) -> str:
    code = str(country_code) if not pd.isna(country_code) else ""
    if code in FIXED_COUNTRY_PRIORITY:
        return code
    if code in EP_MEMBER_OR_VALIDATION_COUNTRY_CODES:
        return "EP member/validation country"
    return "Remaining country"


def create_rd_centers_summary(data: pd.DataFrame) -> pd.DataFrame:
    """Create R&D center summary sorted by patent count for charts."""
    summary = (
        data.groupby(["Priority Country/Region Full"], dropna=False)
        .size()
        .reset_index(name="Patents")
        .rename(columns={"Priority Country/Region Full": "Country"})
    )
    return summary.sort_values(["Patents", "Country"], ascending=[False, True]).reset_index(drop=True)


def create_years_summary(data: pd.DataFrame) -> pd.DataFrame:
    current_year = pd.Timestamp.today().year
    start_year = current_year - PUBLICATION_LOOKBACK_YEARS
    year_index = pd.Index(range(start_year, current_year + 1), name="Years")

    application_counts = (
        data.loc[data["Application Year"].between(start_year, current_year, inclusive="both"), "Application Year"]
        .value_counts()
        .reindex(year_index, fill_value=0)
        .sort_index(ascending=False)
    )
    granted_publication_counts = (
        data.loc[
            data["Publication Year"].between(start_year, current_year, inclusive="both")
            & data["Application Status"].eq("Granted"),
            "Publication Year",
        ]
        .value_counts()
        .reindex(year_index, fill_value=0)
        .sort_index(ascending=False)
    )

    return pd.DataFrame(
        {
            "Years": application_counts.index.astype(int),
            "Application Year Count": application_counts.astype(int).values,
            "Granted Publication Year Count": granted_publication_counts.astype(int).values,
        }
    )


def create_legal_status_summary(data: pd.DataFrame) -> pd.DataFrame:
    """Create a Legal Status summary using Active/Inactive rows from Total Records."""
    if "Dead_Alive" not in data.columns:
        return pd.DataFrame(columns=["Legal Status", "Count"])

    summary = (
        data.loc[data["Dead_Alive"].isin(["Active", "Inactive"]), "Dead_Alive"]
        .value_counts()
        .reindex(["Active", "Inactive"], fill_value=0)
        .rename_axis("Legal Status")
        .reset_index(name="Count")
    )
    return summary


def create_validation_sheet(data: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Publication Number",
        "Application Number",
        "Application Status",
        "Simple Legal Status",
        "Dead_Alive",
        "Dead_Alive Source",
        "Patent Type",
        "Patent Type Source",
        "Priority Date",
        "Link",
    ]
    validation = data[[column for column in columns if column in data.columns]].copy()
    validation["PatSnap Current Legal Status"] = ""
    validation["Validated Dead_Alive"] = ""
    validation["Validation Notes"] = ""
    return validation


def build_column_order(data: pd.DataFrame) -> list[str]:
    preferred = [
        "Publication Number",
        "Country Code",
        "Country Name",
        "Title",
        "Application Number",
        "Application Country Code",
        "Filing Date",
        "Application Year",
        "Publication Date",
        "Publication Year",
        "Priority Date",
        "CPCs",
        "Simple Legal Status",
        "ProjectX Legal Status",
        "Dead_Alive",
        "Dead_Alive Source",
        "Patent Type",
        "Patent Type Source",
        "Application Status Code",
        "Application Status",
        "A+R",
        "A+R Relevancy",
        "Legal Status",
        "Forward Citation Unique Patent Numbers",
        "Count of Cited by Patents",
        "Count of Cited by Patents Source",
        "INPADOC Family ID",
        "INPADOC Family Members (Beta)",
        "Simple Family ID",
        "Priority Number",
        "Priority Country Code Raw",
        "Priority Country Code",
        "Priority Country/Region",
        "Priority Country/Region Full",
        "Current Assignee",
        "Current Assignees",
        "Relevancy Checking",
        "Link",
        "Applicants",
        "Original Assignees",
        "Inventors",
        "Country Priority Rank",
        "Application Status Sort Rank",
        "Dead_Alive Sort Rank",
    ]
    ordered = [column for column in preferred if column in data.columns]
    remaining = [column for column in data.columns if column not in ordered]
    return ordered + remaining


def make_user_facing_output(data: pd.DataFrame) -> pd.DataFrame:
    """Remove internal helper/audit columns from user-facing output sheets."""
    columns_to_drop = [column for column in USER_FACING_COLUMNS_TO_HIDE if column in data.columns]
    return data.drop(columns=columns_to_drop).copy()


def make_process_file_output(data: pd.DataFrame) -> pd.DataFrame:
    """Keep the internal audit sheet, but omit the requested columns from process file."""
    columns_to_drop = [
        column
        for column in ["A+R", "A+R Relevancy", "Application Country Code"]
        if column in data.columns
    ]
    output = data.drop(columns=columns_to_drop).copy()
    output = output.rename(columns={"Current Assignees": "Current Assignee after A+R"})

    # Export dates without time components for the process file sheet only.
    for column in output.columns:
        if pd.api.types.is_datetime64_any_dtype(output[column]):
            output[column] = output[column].dt.date

    return output


# ---------------------------------------------------------------------------
# Excel output helpers
# ---------------------------------------------------------------------------

def fill_output_blanks(data: pd.DataFrame) -> pd.DataFrame:
    """Fill all blank/NA cells with '-' for final Excel output only."""
    output = data.copy()
    output = output.fillna("-")
    output = output.replace(r"^\s*$", "-", regex=True)
    return output


def autosize_worksheets(writer: pd.ExcelWriter) -> None:
    workbook = writer.book
    for worksheet in workbook.worksheets:
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells[:200]:
                value = cell.value
                if value is not None:
                    max_length = max(max_length, len(str(value)))
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)
        worksheet.freeze_panes = "A2"


def add_bar_chart(worksheet: Any, title: str, value_column: int, category_column: int = 1, anchor: str = "G2", max_rows: int = 20) -> None:
    if BarChart is None or Reference is None or worksheet.max_row < 2:
        return
    last_row = min(worksheet.max_row, max_rows + 1)
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = "Patents"
    chart.x_axis.title = "Category"
    values = Reference(worksheet, min_col=value_column, min_row=1, max_row=last_row)
    categories = Reference(worksheet, min_col=category_column, min_row=2, max_row=last_row)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 7
    chart.width = 12
    worksheet.add_chart(chart, anchor)


def add_years_chart(worksheet: Any) -> None:
    if LineChart is None or Reference is None or worksheet.max_row < 2:
        return
    chart = LineChart()
    chart.title = "Application and Granted Publication Trends"
    chart.y_axis.title = "Patents"
    chart.x_axis.title = "Year"
    values = Reference(worksheet, min_col=2, max_col=3, min_row=1, max_row=worksheet.max_row)
    categories = Reference(worksheet, min_col=1, min_row=2, max_row=worksheet.max_row)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 7
    chart.width = 14
    worksheet.add_chart(chart, "E2")


def write_output(
    output_path: Path,
    raw_backup: pd.DataFrame,
    process_file: pd.DataFrame,
    total_records: pd.DataFrame,
    unique_applications: pd.DataFrame,
    unique_families: pd.DataFrame,
    country_summary: pd.DataFrame,
    rd_centers: pd.DataFrame,
    years_summary: pd.DataFrame,
    legal_status_summary: pd.DataFrame,
    validation_sheet: pd.DataFrame,
    dedupe_summary: pd.DataFrame,
    mapping_issues: pd.DataFrame,
    run_notes: pd.DataFrame,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl", datetime_format="yyyy-mm-dd", date_format="yyyy-mm-dd") as writer:
        fill_output_blanks(raw_backup).to_excel(writer, index=False, sheet_name="Raw Backup")
        fill_output_blanks(process_file).to_excel(writer, index=False, sheet_name="process file")
        fill_output_blanks(unique_applications).to_excel(writer, index=False, sheet_name="Total Records")
        fill_output_blanks(unique_families).to_excel(writer, index=False, sheet_name=UNIQUE_FAMILY_SHEET_NAME)
        fill_output_blanks(country_summary).to_excel(writer, index=False, sheet_name="Country")
        fill_output_blanks(rd_centers).to_excel(writer, index=False, sheet_name="R&D Centers")
        fill_output_blanks(years_summary).to_excel(writer, index=False, sheet_name="Years")
        fill_output_blanks(legal_status_summary).to_excel(writer, index=False, sheet_name="Legal Status")
        autosize_worksheets(writer)

        if "Country" in writer.sheets:
            add_bar_chart(writer.sheets["Country"], "Top Countries", value_column=2, category_column=1)
        if "R&D Centers" in writer.sheets:
            add_bar_chart(writer.sheets["R&D Centers"], "Top R&D Centers", value_column=2, category_column=1)
        if "Years" in writer.sheets:
            add_years_chart(writer.sheets["Years"])


# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------
def process_portfolio(input_path: Path, output_path: Path, sheet_name: str) -> None:
    notes: list[dict[str, Any]] = []
    status_map, mapping_issues = build_application_status_map()

    raw_data = pd.read_excel(input_path, sheet_name=sheet_name, dtype=object)
    raw_backup = raw_data.copy()
    add_note(notes, "Input", f"Read sheet '{sheet_name}' from {input_path}", len(raw_data))

    data = raw_data.copy()
    data = validate_and_add_optional_columns(data, notes)
    data = clean_text_columns(data)
    data["Current Assignee"] = data["Current Assignees"]
    data = fill_current_assignee(data, notes)
    data = prepare_identifiers(data)
    data = prepare_dates_and_years(data)
    data = apply_publication_date_filter(data, notes)
    data = add_country_fields(data)
    data = add_application_status(data, status_map, notes)
    data = add_patent_type(data, notes)
    data = create_ar_and_relevancy_fields(data)
    data = add_dead_alive(data)
    data = add_forward_citation_counts(data, notes)

    if EXCLUDE_PROVISIONAL_RECORDS:
        before = len(data)
        data = data.loc[~data["Application Status"].eq("Provisional")].copy()
        add_note(notes, "Provisional filter", "Removed provisional records", before - len(data))
    else:
        add_note(notes, "Provisional filter", "Skipped; provisional records retained", int(data["Application Status"].eq("Provisional").sum()))

    data, country_rank_map = build_country_priority(data)
    data = add_sort_columns(data)

    total_records_internal = sort_for_application_dedupe(data)
    total_records_internal = total_records_internal[build_column_order(total_records_internal)]

    # Full internal dataset for audit/debugging, with the requested columns removed.
    process_file = make_process_file_output(total_records_internal)

    unique_applications_internal = create_unique_applications(total_records_internal, notes)
    unique_applications_internal = sort_for_family_dedupe(unique_applications_internal)

    unique_families_internal, dedupe_summary = create_unique_patent_families(unique_applications_internal)
    unique_families_internal = unique_families_internal[build_column_order(unique_families_internal)]

    country_summary = create_country_summary(total_records_internal, country_rank_map)
    rd_centers = create_rd_centers_summary(unique_families_internal)
    years_summary = create_years_summary(total_records_internal)
    legal_status_summary = create_legal_status_summary(total_records)
    validation_sheet = make_user_facing_output(create_validation_sheet(unique_families_internal))

    # User-facing data sheets exclude internal helper/audit columns.
    total_records = make_user_facing_output(total_records_internal)
    unique_applications = make_user_facing_output(unique_applications_internal)
    unique_families = make_user_facing_output(unique_families_internal)

    if mapping_issues.empty:
        mapping_issues = pd.DataFrame([{"Message": "No duplicate application-status mapping keys detected."}])

    add_note(notes, "Output", "process file", len(total_records_internal))
    add_note(notes, "Output", "Total Records", len(unique_applications))
    add_note(notes, "Output", UNIQUE_FAMILY_SHEET_NAME, len(unique_families))
    add_note(notes, "Manual validation", "PatSnap validation remains a manual review step.")

    run_notes = pd.DataFrame(notes)
    write_output(
        output_path=output_path,
        raw_backup=raw_backup,
        process_file=process_file,
        total_records=total_records,
        unique_applications=unique_applications,
        unique_families=unique_families,
        country_summary=country_summary,
        rd_centers=rd_centers,
        years_summary=years_summary,
        legal_status_summary=legal_status_summary,
        validation_sheet=validation_sheet,
        dedupe_summary=dedupe_summary,
        mapping_issues=mapping_issues,
        run_notes=run_notes,
    )
